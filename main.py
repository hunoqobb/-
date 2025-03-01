import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import csv
from datetime import datetime
from pypinyin import pinyin, Style

# 定义帮助文本
help_text = """使用说明：

1. 搜索功能：
   - 在左上方输入框可以按标题、作者或朝代搜索诗词
   - 支持模糊搜索，输入部分内容即可
   - 按回车键或点击搜索按钮进行搜索

2. 诗词列表：
   - 左侧显示所有诗词的列表
   - 点击任意诗词可在右侧查看详细内容

3. 诗词详情：
   - 右侧上方显示诗词原文和拼音
   - 下方标签页包含译文、注释、赏析和作者介绍

4. 编辑功能：
   - 点击"编辑"按钮可以修改诗词内容
   - 修改完成后点击"保存"按钮保存更改
   - 点击"取消"可放弃更改

5. 导入导出：
   - 可以导入其他JSON格式的诗词数据
   - 可以导出当前所有诗词数据为JSON文件"""

# 定义关于文本
about_text = """少儿古诗词学习 v1.0

本软件旨在帮助儿童学习中国古诗词，提供了以下功能：
- 诗词原文与拼音对照
- 译文与注释辅助理解
- 诗词赏析提升鉴赏能力
- 作者介绍了解历史背景

特点：
- 简洁直观的界面设计
- 方便的搜索和编辑功能
- 支持数据导入导出

作者：奋青
联系方式：393283@qq.com

免责声明：
本软件为开源项目，按"原样"提供，不提供任何明示或暗示的保证。作者不对使用本软件造成的任何直接或间接损失负责。用户可以自由使用、修改和分发本软件，但必须遵守相关开源协议的规定。

版权所有 © 2025"""

class PoemApp:
    def __init__(self, root):
        self.root = root
        self.root.title('少儿古诗词学习')
        self.root.iconbitmap('logo.ico')
        self.editing = False
        
        # 创建style对象
        style = ttk.Style()
        
        # 绑定窗口关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self.close_app)
        
        # 使用更大的窗口尺寸
        width = 1400
        height = 800
        
        # 获取屏幕尺寸
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        
        # 计算窗口位置使其居中
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        # 设置窗口大小和位置
        self.root.geometry(f'{width}x{height}+{x}+{y}')

        # 创建标题栏
        title_frame = ttk.Frame(root)
        title_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)
        
        # 创建一个内部框架来容纳标题和图片
        title_container = ttk.Frame(title_frame)
        title_container.pack(side=tk.TOP, pady=10, expand=True)
        
        # 加载并显示图片
        logo_image = tk.PhotoImage(file='logo.png')
        logo_label = ttk.Label(title_container, image=logo_image)
        logo_label.image = logo_image  # 保持引用防止被垃圾回收
        logo_label.pack(side=tk.LEFT, padx=10)
        
        # 系统标题居中显示，使用更大字体
        self.title_label = ttk.Label(title_container, text='少儿古诗词学习', font=('SimSun', 28, 'bold'))
        self.title_label.pack(side=tk.LEFT, padx=10)

        # 创建菜单栏
        menubar = tk.Menu(root)
        root.config(menu=menubar)
        
        # 创建帮助菜单
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="帮助", menu=help_menu)
        help_menu.add_command(label="帮助", command=lambda: self.show_help_dialog())
        help_menu.add_command(label="关于", command=lambda: self.show_about_dialog())
        
        # 标题居中显示
        self.title_label.pack(side=tk.TOP, pady=10, expand=True)

        # 加载诗词数据
        with open('poems.json', 'r', encoding='utf-8') as f:
            self.poems_data = json.load(f)['poems']

        # 创建主框架
        self.main_frame = ttk.Frame(root, padding="10")
        self.main_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(1, weight=1)
        self.main_frame.grid_columnconfigure(1, weight=1)
        self.main_frame.grid_rowconfigure(0, weight=1)

        # 创建左右分栏
        self.left_frame = ttk.Frame(self.main_frame)
        self.left_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5))
        
        self.right_frame = ttk.Frame(self.main_frame)
        self.right_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))

        # 创建搜索框架
        self.search_frame = ttk.Frame(self.left_frame)
        self.search_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))

        # 创建搜索条件输入框
        ttk.Label(self.search_frame, text="标题:").grid(row=0, column=0, padx=5)
        self.title_search = ttk.Entry(self.search_frame, width=8)
        self.title_search.grid(row=0, column=1, padx=5)
        self.title_search.bind('<Return>', lambda e: self.search_poems())

        ttk.Label(self.search_frame, text="作者:").grid(row=0, column=2, padx=5)
        self.author_search = ttk.Entry(self.search_frame, width=5)
        self.author_search.grid(row=0, column=3, padx=5)
        self.author_search.bind('<Return>', lambda e: self.search_poems())

        ttk.Label(self.search_frame, text="朝代:").grid(row=0, column=4, padx=5)
        self.dynasty_search = ttk.Entry(self.search_frame, width=3)
        self.dynasty_search.grid(row=0, column=5, padx=5)
        self.dynasty_search.bind('<Return>', lambda e: self.search_poems())

        self.search_btn = ttk.Button(self.search_frame, text="搜索", command=self.search_poems, width=4)
        self.search_btn.grid(row=0, column=6, padx=5)
        
        # 添加收藏夹按钮
        self.show_favorites_btn = ttk.Button(self.search_frame, text="收藏夹", command=self.show_favorites, width=6)
        self.show_favorites_btn.grid(row=0, column=7, padx=5)
        
        # 添加显示全部按钮
        self.show_all_btn = ttk.Button(self.search_frame, text="全部", command=self.show_all_poems, width=4)
        self.show_all_btn.grid(row=0, column=8, padx=5)
        
        # 添加批量删除按钮
        self.batch_delete_btn = ttk.Button(self.search_frame, text="批量删除", command=self.batch_delete_poems, width=8)
        self.batch_delete_btn.grid(row=0, column=9, padx=5)
        
        # 初始化批量删除模式标志
        self.batch_delete_mode = False

        # 创建诗词列表的树形视图
        self.poem_tree = ttk.Treeview(self.left_frame, columns=('title', 'author', 'dynasty'), show='headings', selectmode='extended')
        self.poem_tree.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # 设置列标题的样式和背景色
        style = ttk.Style()
        style.configure('Treeview.Heading', background='#E8E8E8', relief='flat', font=('SimSun', 10))
        
        # 设置列标题
        self.poem_tree.heading('title', text='标题', command=lambda: self.sort_tree('title'))
        self.poem_tree.heading('author', text='作者', command=lambda: self.sort_tree('author'))
        self.poem_tree.heading('dynasty', text='朝代', command=lambda: self.sort_tree('dynasty'))
        
        # 修改排序状态记录
        self.sort_column = None  # 当前排序的列
        self.sort_reverse = False  # 是否逆序
        self.sort_count = 0  # 添加点击次数记录
        
        # 配置left_frame的网格权重，使poem_tree能够自动扩展
        self.left_frame.grid_columnconfigure(0, weight=1)
        self.left_frame.grid_rowconfigure(1, weight=1)

        # 设置列宽
        self.poem_tree.column('title', width=150)
        self.poem_tree.column('author', width=100)
        self.poem_tree.column('dynasty', width=40)

        # 绑定选择事件
        self.poem_tree.bind('<<TreeviewSelect>>', self.show_poem_details)

        # 添加滚动条
        tree_scroll = ttk.Scrollbar(self.left_frame, orient=tk.VERTICAL, command=self.poem_tree.yview)
        tree_scroll.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.poem_tree.configure(yscrollcommand=tree_scroll.set)

        # 创建右侧内容框架
        self.content_frame = ttk.Frame(self.right_frame)
        self.content_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.content_frame.grid_columnconfigure(0, weight=1)
        self.content_frame.grid_rowconfigure(1, weight=2)  # 诗词显示框占2份
        self.content_frame.grid_rowconfigure(2, weight=3)  # 注释内容显示框占3份

        # 创建右侧顶部按钮框架
        self.right_top_frame = ttk.Frame(self.content_frame)
        self.right_top_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        self.right_top_frame.grid_columnconfigure(0, weight=1)

        # 创建一个内部框架来容纳按钮，实现居中效果
        button_container = ttk.Frame(self.right_top_frame)
        button_container.grid(row=0, column=0)

        # 添加编辑、保存、取消、关闭、导入导出按钮到按钮容器中
        self.add_btn = ttk.Button(button_container, text='添加', command=self.add_poem)
        self.add_btn.pack(side=tk.LEFT, padx=5)
        self.favorite_btn = ttk.Button(button_container, text='收藏', command=self.toggle_favorite)
        self.favorite_btn.pack(side=tk.LEFT, padx=5)
        self.edit_btn = ttk.Button(button_container, text='编辑', command=self.edit_poem)
        self.edit_btn.pack(side=tk.LEFT, padx=5)
        self.save_btn = ttk.Button(button_container, text='保存', command=self.save_poem, state='disabled')
        self.save_btn.pack(side=tk.LEFT, padx=5)
        self.cancel_btn = ttk.Button(button_container, text='取消', command=self.cancel_edit, state='disabled')
        self.cancel_btn.pack(side=tk.LEFT, padx=5)
        self.import_btn = ttk.Button(button_container, text='导入', command=self.import_poems)
        self.import_btn.pack(side=tk.LEFT, padx=5)
        self.export_btn = ttk.Button(button_container, text='导出', command=self.export_poems)
        self.export_btn.pack(side=tk.LEFT, padx=5)
        self.close_btn = ttk.Button(button_container, text='退出', command=self.close_app)
        self.close_btn.pack(side=tk.LEFT, padx=5)

        # 修改诗词内容文本框的创建
        self.poem_content = tk.Text(self.content_frame, wrap=tk.WORD, font=('SimSun', 14), 
                                   relief=tk.SUNKEN, borderwidth=1)
        self.poem_content.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 2))
        
        # 添加滚动条
        content_scroll = ttk.Scrollbar(self.content_frame, orient=tk.VERTICAL, command=self.poem_content.yview)
        content_scroll.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.poem_content.configure(yscrollcommand=content_scroll.set)
        
        # 创建朗读控制面板，减小padding值使其更紧凑
        self.read_control_frame = ttk.LabelFrame(self.content_frame, text='朗读控制', padding=2)
        self.read_control_frame.grid(row=1, column=2, sticky=(tk.N, tk.S, tk.E), padx=5, pady=5)
        
        # 添加音色选择，减小间距
        ttk.Label(self.read_control_frame, text='音色:').pack(pady=(2,0))
        self.voice_var = tk.StringVar()
        self.voice_combo = ttk.Combobox(self.read_control_frame, textvariable=self.voice_var, state='readonly')
        self.voice_combo.pack(fill=tk.X, pady=1)
        
        # 添加朗读按钮，减小间距
        self.read_btn = ttk.Button(self.read_control_frame, text='朗读', command=self.read_poem)
        self.read_btn.pack(fill=tk.X, pady=1)
        
        # 添加暂停/继续按钮
        self.pause_btn = ttk.Button(self.read_control_frame, text='暂停', command=self.pause_resume_reading, state='disabled')
        self.pause_btn.pack(fill=tk.X, pady=1)
        
        # 添加停止按钮
        self.stop_btn = ttk.Button(self.read_control_frame, text='停止', command=self.stop_reading, state='disabled')
        self.stop_btn.pack(fill=tk.X, pady=1)
        
        # 添加语速控制，减小间距
        ttk.Label(self.read_control_frame, text='语速:').pack(pady=(5,0))
        self.rate_scale = ttk.Scale(self.read_control_frame, from_=50, to=300, orient=tk.HORIZONTAL)
        self.rate_scale.set(150)
        self.rate_scale.pack(fill=tk.X, pady=1)
        
        # 添加音量控制
        ttk.Label(self.read_control_frame, text='音量:').pack(pady=(5,0))
        self.volume_scale = ttk.Scale(self.read_control_frame, from_=0, to=1, orient=tk.HORIZONTAL)
        self.volume_scale.set(1.0)
        self.volume_scale.pack(fill=tk.X, pady=1)
        
        # 添加退出按钮到朗读控制区域下方
        self.close_btn = ttk.Button(self.read_control_frame, text='退出程序', command=self.close_app, width=8)
        self.close_btn.pack(fill=tk.X, pady=(10,1))
        
        # 创建诗词详情的框架
        self.detail_frame = ttk.Frame(self.content_frame)
        self.detail_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        self.detail_frame.grid_columnconfigure(0, weight=1)
        self.detail_frame.grid_rowconfigure(0, weight=1)
        
        # 创建诗词详情的文本框和标签
        info_frame = ttk.Frame(self.detail_frame)
        info_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        info_frame.grid_columnconfigure(0, weight=1)
        info_frame.grid_rowconfigure(0, weight=1)
        
        # 创建Notebook用于显示不同类型的信息
        self.notebook = ttk.Notebook(info_frame)
        self.notebook.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 创建各个标签页
        translation_frame = ttk.Frame(self.notebook, padding=10)
        note_frame = ttk.Frame(self.notebook, padding=10)
        appreciation_frame = ttk.Frame(self.notebook, padding=10)
        author_frame = ttk.Frame(self.notebook, padding=10)
        
        # 设置每个frame的最小高度和扩展性
        min_height = 200
        for frame in [translation_frame, note_frame, appreciation_frame, author_frame]:
            frame.grid_propagate(False)
            frame.grid_rowconfigure(0, weight=1)
            frame.grid_columnconfigure(0, weight=1)
            frame.configure(height=min_height)
        
        self.notebook.add(translation_frame, text='译文')
        self.notebook.add(note_frame, text='注释')
        self.notebook.add(appreciation_frame, text='赏析')
        self.notebook.add(author_frame, text='作者介绍')
        
        # 为每个标签页创建文本框和滚动条
        for frame, text_attr in [
            (translation_frame, 'translation_text'),
            (note_frame, 'note_text'),
            (appreciation_frame, 'appreciation_text'),
            (author_frame, 'author_text')
        ]:
            container = ttk.Frame(frame)
            container.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.E, tk.W))
            container.grid_columnconfigure(0, weight=1)
            container.grid_rowconfigure(0, weight=1)
            
            text_widget = tk.Text(container, wrap=tk.WORD, 
                                font=('SimSun', 12), spacing1=10, spacing2=5, spacing3=10)
            text_widget.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.E, tk.W))
            
            scroll = ttk.Scrollbar(container, orient=tk.VERTICAL, command=text_widget.yview)
            scroll.grid(row=0, column=1, sticky=(tk.N, tk.S), padx=(0, 2))
            
            text_widget.configure(yscrollcommand=scroll.set)
            setattr(self, text_attr, text_widget)

        # 设置文本框初始状态为禁用
        for text_widget in [self.translation_text, self.note_text, self.appreciation_text, self.author_text]:
            text_widget.config(state='disabled')

        # 配置网格权重
        translation_frame.grid_columnconfigure(0, weight=1)
        translation_frame.grid_rowconfigure(0, weight=1)
        note_frame.grid_columnconfigure(0, weight=1)
        note_frame.grid_rowconfigure(0, weight=1)
        appreciation_frame.grid_columnconfigure(0, weight=1)
        appreciation_frame.grid_rowconfigure(0, weight=1)
        author_frame.grid_columnconfigure(0, weight=1)
        author_frame.grid_rowconfigure(0, weight=1)

        # 配置主窗口的grid权重，使其能够扩展
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(1, weight=1)  # row 1 是main_frame
        
        # 配置main_frame的grid权重
        self.main_frame.grid_columnconfigure(1, weight=3)  # 右侧占更多空间
        self.main_frame.grid_rowconfigure(0, weight=1)
        
        # 配置right_frame的grid权重
        self.right_frame.grid_columnconfigure(0, weight=1)
        self.right_frame.grid_rowconfigure(0, weight=1)
        
        # 配置content_frame的grid权重
        self.content_frame.grid_columnconfigure(0, weight=1)
        self.content_frame.grid_rowconfigure(1, weight=2)  # 诗词显示框占2份
        self.content_frame.grid_rowconfigure(2, weight=3)  # 注释内容显示框占3份

        # 初始化朗读引擎和状态
        self.engine = None
        self.is_reading = False
        self.is_paused = False
        
        # 初始化音色列表
        self.init_voice_list()
        
        self.poem_content.config(state='disabled')

        # 初始化收藏列表
        self.favorites = self.load_favorites()

    def init_voice_list(self):
        try:
            try:
                import pyttsx3
            except ImportError:
                messagebox.showerror('错误', '未找到pyttsx3模块，朗读功能将不可用。\n请使用以下命令安装：\npip install pyttsx3\npip install pywin32')
                self.voice_combo['values'] = ['未安装语音引擎']
                self.voice_combo.set('未安装语音引擎')
                self.read_btn.config(state='disabled')
                return

            if not self.engine:
                self.engine = pyttsx3.init()
            
            # 获取可用的音色列表
            voices = self.engine.getProperty('voices')
            self.voices = voices
            
            # 更新下拉框选项
            voice_names = []
            for voice in voices:
                # 提取音色名称,去掉Microsoft和其他前缀
                name = voice.name
                if 'Microsoft' in name:
                    name = name.replace('Microsoft ', '')
                voice_names.append(name)
            
            if not voice_names:
                messagebox.showwarning('警告', '未检测到系统语音引擎，朗读功能可能无法正常使用。\n请确保系统安装了语音引擎。')
                voice_names = ['系统默认']
            
            self.voice_combo['values'] = voice_names
            self.voice_combo.set(voice_names[0])
                
        except Exception as e:
            messagebox.showerror('错误', f'初始化音色列表失败: {str(e)}\n朗读功能将不可用。')
            self.voice_combo['values'] = ['初始化失败']
            self.voice_combo.set('初始化失败')
            self.read_btn.config(state='disabled')

    def search_poems(self):
        # 获取搜索条件
        title = self.title_search.get().strip()
        author = self.author_search.get().strip()
        dynasty = self.dynasty_search.get().strip()

        # 根据条件筛选诗词
        filtered_poems = self.poems_data
        if title:
            filtered_poems = [poem for poem in filtered_poems if title.lower() in poem['title'].lower()]
        if author:
            filtered_poems = [poem for poem in filtered_poems if author.lower() in poem['author'].lower()]
        if dynasty:
            filtered_poems = [poem for poem in filtered_poems if dynasty.lower() in poem['dynasty'].lower()]

        # 清空现有搜索结果
        for item in self.poem_tree.get_children():
            self.poem_tree.delete(item)

        # 显示搜索结果
        for poem in filtered_poems:
            self.poem_tree.insert('', 'end', values=(poem['title'], poem['author'], poem['dynasty']))
        
        # 如果之前有排序，保持相同的排序
        if self.sort_column:
            self.sort_tree(self.sort_column)

    def show_poem_details(self, event):
        # 获取选中的诗词
        selection = self.poem_tree.selection()
        if not selection:
            return
            
        item = selection[0]
        values = self.poem_tree.item(item)['values']
        title = values[0]
        
        # 查找诗词数据
        poem = None
        for p in self.poems_data:
            if p['title'] == title:
                poem = p
                break
                
        if not poem:
            return
            
        # 处理诗词内容和拼音
        self.poem_content.config(state='normal')
        self.poem_content.delete('1.0', 'end')
        
        # 配置标题、拼音和汉字的样式
        self.poem_content.tag_configure('title', font=('SimSun', 18, 'bold'), justify='center', spacing1=5, spacing3=5)
        self.poem_content.tag_configure('pinyin', font=('SimSun', 12), justify='center', spacing1=2, spacing3=2, spacing2=5)
        self.poem_content.tag_configure('hanzi', font=('SimSun', 16, 'bold'), justify='center', spacing1=2, spacing3=5, spacing2=5)
        
        # 插入标题和拼音
        if 'title_pinyin' in poem:
            self.poem_content.insert('end', poem.get('title_pinyin', '') + '\n', 'pinyin')
        self.poem_content.insert('end', poem['title'] + '\n\n', 'title')
        
        # 获取诗词内容和拼音
        content_lines = poem.get('content', [])
        pinyin_lines = poem.get('content_pinyin', [])
        
        # 将拼音显示在汉字上方
        if isinstance(content_lines, list) and isinstance(pinyin_lines, list) and len(content_lines) == len(pinyin_lines):
            for content_line, pinyin_line in zip(content_lines, pinyin_lines):
                # 插入拼音行
                self.poem_content.insert('end', pinyin_line + '\n', 'pinyin')
                # 插入汉字行
                self.poem_content.insert('end', content_line + '\n', 'hanzi')
        else:
            # 如果没有拼音或格式不匹配，只显示内容
            if isinstance(content_lines, list):
                for line in content_lines:
                    self.poem_content.insert('end', line + '\n', 'hanzi')
            else:
                self.poem_content.insert('end', str(content_lines), 'hanzi')
        
        self.poem_content.config(state='disabled')
        
        # 显示诗词内容
        for text_widget in [self.translation_text, self.note_text, self.appreciation_text, self.author_text]:
            text_widget.config(state='normal')
            text_widget.delete('1.0', 'end')
            
        self.translation_text.insert('1.0', poem.get('translation', ''))
        self.note_text.insert('1.0', poem.get('note', ''))
        self.appreciation_text.insert('1.0', poem.get('appreciation', ''))
        self.author_text.insert('1.0', poem.get('author_intro', ''))
        
        # 如果不在编辑模式，禁用文本框
        if not self.editing:
            for text_widget in [self.translation_text, self.note_text, self.appreciation_text, self.author_text]:
                text_widget.config(state='disabled')

        # 更新收藏按钮状态
        poem_id = f"{poem['title']}_{poem['author']}"
        if poem_id in self.favorites:
            self.favorite_btn.config(text='取消收藏')
        else:
            self.favorite_btn.config(text='收藏')

    def edit_poem(self):
        # 启用编辑模式
        self.editing = True
        self.edit_btn.config(state='disabled')
        self.save_btn.config(state='normal')
        self.cancel_btn.config(state='normal')
        
        # 启用所有文本框的编辑，包括诗词内容
        self.poem_content.config(state='normal')
        for text_widget in [self.translation_text, self.note_text, self.appreciation_text, self.author_text]:
            text_widget.config(state='normal')

    def save_poem(self):
        # 获取当前选中的诗词
        selection = self.poem_tree.selection()
        if not selection:
            return
            
        item = selection[0]
        values = self.poem_tree.item(item)['values']
        title = values[0]
        
        # 获取诗词内容
        content_text = self.poem_content.get('1.0', 'end-1c')
        lines = content_text.split('\n')
        
        # 解析标题、拼音和内容
        title_pinyin = ''
        new_title = ''
        content_lines = []
        content_pinyin_lines = []
        
        # 处理标题部分
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if not line:  # 跳过空行
                i += 1
                continue
            if not title_pinyin and not new_title:  # 第一个非空行是标题拼音
                title_pinyin = line
                i += 1
            elif not new_title:  # 第二个非空行是标题
                new_title = line
                i += 2  # 跳过标题后的空行
                break
            else:
                break
        
        # 处理内容部分
        while i < len(lines):
            line = lines[i].strip()
            if line:  # 非空行
                if len(content_lines) == len(content_pinyin_lines):  # 当前是拼音行
                    content_pinyin_lines.append(line)
                else:  # 当前是汉字行
                    content_lines.append(line)
            i += 1
        
        # 更新诗词数据
        for poem in self.poems_data:
            if poem['title'] == title:
                poem['title'] = new_title
                poem['title_pinyin'] = title_pinyin
                poem['content'] = content_lines
                poem['content_pinyin'] = content_pinyin_lines
                poem['translation'] = self.translation_text.get('1.0', 'end-1c')
                poem['note'] = self.note_text.get('1.0', 'end-1c')
                poem['appreciation'] = self.appreciation_text.get('1.0', 'end-1c')
                poem['author_intro'] = self.author_text.get('1.0', 'end-1c')
                break
                
        # 保存到文件
        with open('poems.json', 'w', encoding='utf-8') as f:
            json.dump({'poems': self.poems_data}, f, ensure_ascii=False, indent=4)
            
        # 禁用编辑模式
        self.cancel_edit()
        
        # 更新树形视图中的标题
        self.poem_tree.item(item, values=(new_title, values[1], values[2]))
        
        messagebox.showinfo('成功', '保存成功！')

    def cancel_edit(self):
        # 禁用编辑模式
        self.editing = False
        self.edit_btn.config(state='normal')
        self.save_btn.config(state='disabled')
        self.cancel_btn.config(state='disabled')
        
        # 禁用所有文本框的编辑
        for text_widget in [self.translation_text, self.note_text, self.appreciation_text, self.author_text]:
            text_widget.config(state='disabled')
            
        # 重新显示当前诗词
        self.show_poem_details(None)

    def read_poem(self):
        # 检查是否已安装pyttsx3
        try:
            import pyttsx3
        except ImportError:
            messagebox.showerror('错误', '未找到pyttsx3模块，无法使用朗读功能。\n请使用以下命令安装：\npip install pyttsx3\npip install pywin32')
            return
        
        # 获取当前选中的诗词内容
        selection = self.poem_tree.selection()
        if not selection:
            messagebox.showinfo('提示', '请先选择一首诗词')
            return
            
        # 获取诗词内容
        content = ''
        for line in self.poem_content.get('1.0', 'end-1c').split('\n'):
            if line.strip() and not any(char.isascii() for char in line):  # 只读取汉字行
                content += line.strip() + '，'
                
        if not content:
            messagebox.showinfo('提示', '无法获取诗词内容')
            return
            
        try:
            import threading
            
            # 初始化引擎
            if not self.engine:
                self.engine = pyttsx3.init()
            
            # 设置选中的音色
            try:
                voice_name = self.voice_var.get()
                if hasattr(self, 'voices'):
                    for voice in self.voices:
                        if voice.name.replace('Microsoft ', '') == voice_name:
                            self.engine.setProperty('voice', voice.id)
                            break
            except Exception as e:
                print(f"设置音色失败: {str(e)}")  # 仅打印错误，继续使用默认音色
            
            # 更新引擎属性
            self.engine.setProperty('rate', self.rate_scale.get())  # 设置语速
            self.engine.setProperty('volume', self.volume_scale.get())  # 设置音量
            
            # 启用控制按钮
            self.read_btn.config(state='disabled')
            self.pause_btn.config(state='normal')
            self.stop_btn.config(state='normal')
            
            # 开始朗读
            self.is_reading = True
            self.is_paused = False
            
            def read_thread():
                try:
                    # 将内容按句号、逗号等标点符号分段
                    segments = [s.strip() for s in content.replace('，', '。').split('。') if s.strip()]
                    
                    for segment in segments:
                        if not self.is_reading:
                            break
                            
                        while self.is_paused:
                            if not self.is_reading:
                                break
                            import time
                            time.sleep(0.1)
                        
                        if not self.is_reading:
                            break
                            
                        # 更新引擎属性
                        self.engine.setProperty('rate', self.rate_scale.get())
                        self.engine.setProperty('volume', self.volume_scale.get())
                        
                        self.engine.say(segment)
                        self.engine.runAndWait()
                    
                    # 朗读完成后恢复按钮状态
                    if self.is_reading:  # 只有在非手动停止的情况下才执行
                        self.root.after(0, self.reset_read_buttons)
                        
                except Exception as e:
                    self.root.after(0, lambda: messagebox.showerror('错误', f'朗读过程中发生错误: {str(e)}'))
                    self.root.after(0, self.reset_read_buttons)
            
            # 在新线程中执行朗读
            threading.Thread(target=read_thread, daemon=True).start()
            
        except Exception as e:
            messagebox.showerror('错误', f'启动朗读失败: {str(e)}')
            self.reset_read_buttons()
    
    def pause_resume_reading(self):
        if not self.engine or not self.is_reading:
            return
            
        if self.is_paused:
            # 继续朗读
            self.is_paused = False
            self.pause_btn.config(text='暂停')
        else:
            # 暂停朗读
            self.is_paused = True
            self.pause_btn.config(text='继续')
    
    def stop_reading(self):
        if not self.engine:
            return
            
        # 停止朗读
        self.is_reading = False
        self.is_paused = False
        self.engine.stop()
        self.reset_read_buttons()
        
        # 重新初始化引擎，以便下次朗读
        self.engine = None
    
    def reset_read_buttons(self):
        # 重置按钮状态
        self.read_btn.config(state='normal')
        self.pause_btn.config(state='disabled')
        self.stop_btn.config(state='disabled')
        self.is_reading = False
        self.is_paused = False
        if self.pause_btn['text'] == '继续':
            self.pause_btn.config(text='暂停')

    def close_app(self):
        # 弹出确认对话框
        if messagebox.askyesno('确认退出', '确定要退出程序吗？'):
            try:
                # 停止朗读并清理引擎资源
                if self.engine:
                    self.stop_reading()
                    self.engine = None
                
                # 确保所有子窗口都被关闭
                for widget in self.root.winfo_children():
                    if isinstance(widget, tk.Toplevel):
                        widget.destroy()
                
                # 直接退出程序
                import sys
                sys.exit(0)
                
            except Exception as e:
                messagebox.showerror('错误', f'关闭程序时发生错误：{str(e)}')
                return

    def add_poem(self):
        # 创建对话框
        dialog = tk.Toplevel(self.root)
        dialog.title('添加诗词')
        dialog.geometry('600x800')
        
        # 设置对话框在屏幕中间
        root_x = self.root.winfo_x()
        root_y = self.root.winfo_y()
        root_width = self.root.winfo_width()
        root_height = self.root.winfo_height()
        
        dialog_width = 600
        dialog_height = 800
        x = root_x + (root_width - dialog_width) // 2
        y = root_y + (root_height - dialog_height) // 2
        dialog.geometry(f'{dialog_width}x{dialog_height}+{x}+{y}')
        
        # 设置对话框为模态
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 创建主框架，添加内边距
        main_frame = ttk.Frame(dialog, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        dialog.grid_columnconfigure(0, weight=1)
        dialog.grid_rowconfigure(0, weight=1)
        
        # 添加标题标签
        title_label = ttk.Label(main_frame, text="添加新诗词", font=('SimSun', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # 创建输入框架
        input_frame = ttk.Frame(main_frame)
        input_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        main_frame.grid_columnconfigure(0, weight=1)
        
        # 创建标签和输入框，使用网格布局
        labels = ['标题:', '作者:', '朝代:', '内容:', '译文:', '注释:', '赏析:', '作者介绍:']
        entries = {}
        
        for i, label in enumerate(labels):
            # 创建标签框架
            label_frame = ttk.Frame(input_frame)
            label_frame.grid(row=i, column=0, sticky=(tk.W, tk.E), pady=5)
            label_frame.grid_columnconfigure(1, weight=1)
            
            # 添加标签
            ttk.Label(label_frame, text=label, font=('SimSun', 11)).grid(row=0, column=0, padx=(0, 10), sticky=tk.W)
            
            # 根据字段类型创建不同的输入控件
            if label in ['内容:', '译文:', '注释:', '赏析:', '作者介绍:']:
                # 创建文本框和滚动条的容器
                text_container = ttk.Frame(label_frame)
                text_container.grid(row=0, column=1, sticky=(tk.W, tk.E))
                text_container.grid_columnconfigure(0, weight=1)
                
                # 创建文本框
                text = tk.Text(text_container, height=4 if label == '内容:' else 3,
                              wrap=tk.WORD, font=('SimSun', 11))
                text.grid(row=0, column=0, sticky=(tk.W, tk.E))
                
                # 添加滚动条
                scrollbar = ttk.Scrollbar(text_container, orient=tk.VERTICAL, command=text.yview)
                scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
                text.configure(yscrollcommand=scrollbar.set)
                
                entries[label.rstrip(':')] = text
            else:
                # 创建单行输入框
                entry = ttk.Entry(label_frame, font=('SimSun', 11))
                entry.grid(row=0, column=1, sticky=(tk.W, tk.E))
                entries[label.rstrip(':')] = entry
        
        # 创建按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, pady=20)
        
        def save_new_poem():
            # 获取输入内容
            title = entries['标题'].get().strip()
            author = entries['作者'].get().strip()
            dynasty = entries['朝代'].get().strip()
            content = entries['内容'].get('1.0', tk.END).strip().split('\n')
            translation = entries['译文'].get('1.0', tk.END).strip()
            note = entries['注释'].get('1.0', tk.END).strip()
            appreciation = entries['赏析'].get('1.0', tk.END).strip()
            author_intro = entries['作者介绍'].get('1.0', tk.END).strip()
            
            # 验证必填字段
            if not all([title, author, dynasty, content]):
                messagebox.showerror('错误', '标题、作者、朝代和内容为必填项！')
                return
            
            # 生成拼音
            title_pinyin = ' '.join([' '.join(p) for p in pinyin(title, style=Style.TONE)])
            content_pinyin = []
            for line in content:
                if line.strip():  # 跳过空行
                    py = pinyin(line, style=Style.TONE)
                    content_pinyin.append(' '.join([' '.join(p) for p in py]))
            
            # 创建新诗词
            new_poem = {
                'title': title,
                'author': author,
                'dynasty': dynasty,
                'content': content,
                'translation': translation,
                'note': note,
                'appreciation': appreciation,
                'author_intro': author_intro,
                'title_pinyin': title_pinyin,
                'content_pinyin': content_pinyin
            }
            
            # 添加到数据中
            self.poems_data.append(new_poem)
            
            # 保存到文件
            with open('poems.json', 'w', encoding='utf-8') as f:
                json.dump({'poems': self.poems_data}, f, ensure_ascii=False, indent=4)
            
            # 刷新显示
            self.search_poems()
            
            # 关闭窗口
            dialog.destroy()
            messagebox.showinfo('成功', '添加诗词成功！')
        
        # 创建按钮
        save_btn = ttk.Button(button_frame, text='保存', command=save_new_poem)
        save_btn.pack(side=tk.LEFT, padx=10)
        
        cancel_btn = ttk.Button(button_frame, text='取消', command=dialog.destroy)
        cancel_btn.pack(side=tk.LEFT, padx=10)

    def import_poems(self):
        try:
            # 打开文件选择对话框
            file_path = filedialog.askopenfilename(
                title='选择要导入的文件',
                filetypes=[
                    ('Excel files', '*.xlsx'),
                    ('JSON files', '*.json'),
                    ('CSV files', '*.csv')
                ]
            )
            
            if not file_path:
                return
                
            if file_path.lower().endswith('.json'):
                # 读取JSON文件
                with open(file_path, 'r', encoding='utf-8') as f:
                    imported_data = json.load(f)
                    
                if 'poems' not in imported_data:
                    raise ValueError('无效的JSON格式')
                    
                new_poems = imported_data['poems']
                
            elif file_path.lower().endswith('.xlsx'):
                try:
                    import openpyxl
                except ImportError:
                    messagebox.showerror('错误', '未找到openpyxl模块，无法导入Excel文件。\n请使用以下命令安装：\npip install openpyxl')
                    return
                    
                # 读取Excel文件
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                # 获取表头
                headers = [cell.value for cell in ws[1]]
                
                # 读取数据
                new_poems = []
                for row in ws.iter_rows(min_row=2):
                    poem = {}
                    for header, cell in zip(headers, row):
                        if header in ['content', 'content_pinyin']:
                            # 将字符串转换为列表
                            value = cell.value or ''
                            poem[header] = value.split('|') if value else []
                        else:
                            poem[header] = cell.value or ''
                    if poem.get('title'):  # 只添加有标题的诗词
                        new_poems.append(poem)
                    
            else:  # CSV文件
                new_poems = []
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            # 处理content和content_pinyin字段，将字符串转换为列表
                            content = row.get('content', '').split('|')
                            content_pinyin = row.get('content_pinyin', '').split('|')
                            
                            poem = {
                                'title': row.get('title', ''),
                                'author': row.get('author', ''),
                                'dynasty': row.get('dynasty', ''),
                                'content': content,
                                'content_pinyin': content_pinyin,
                                'translation': row.get('translation', ''),
                                'note': row.get('note', ''),
                                'appreciation': row.get('appreciation', ''),
                                'author_intro': row.get('author_intro', ''),
                                'title_pinyin': row.get('title_pinyin', '')
                            }
                            if poem['title']:  # 只添加有标题的诗词
                                new_poems.append(poem)
                except Exception as e:
                    messagebox.showerror('错误', f'导入CSV文件时发生错误：{str(e)}')
                    return
            
            # 为没有拼音的诗词添加拼音
            for poem in new_poems:
                if not poem.get('title_pinyin'):
                    title_py = pinyin(poem['title'], style=Style.TONE)
                    poem['title_pinyin'] = ' '.join([' '.join(p) for p in title_py])
                
                if not any(poem.get('content_pinyin', [])):
                    content_pinyin = []
                    for line in poem.get('content', []):
                        line_py = pinyin(line, style=Style.TONE)
                        content_pinyin.append(' '.join([' '.join(p) for p in line_py]))
                    poem['content_pinyin'] = content_pinyin
                        
            # 处理重复诗词
            duplicate_count = 0
            skip_count = 0
            for new_poem in new_poems[:]:  # 使用切片创建副本以便在循环中修改
                for existing_poem in self.poems_data:
                    if (new_poem['title'] == existing_poem['title'] and 
                        new_poem['author'] == existing_poem['author']):
                        # 找到重复诗词
                        response = messagebox.askyesnocancel(
                            '发现重复诗词',
                            f'发现重复诗词：\n标题：{new_poem["title"]}\n作者：{new_poem["author"]}\n\n'
                            f'是否覆盖已有诗词？\n'
                            f'是 - 覆盖\n否 - 跳过\n取消 - 停止导入'
                        )
                        
                        if response is None:  # 取消导入
                            messagebox.showinfo('提示', 
                                f'导入已取消。\n成功导入：{len(new_poems) - len(new_poems[new_poems.index(new_poem):])}首\n'
                                f'覆盖：{duplicate_count}首\n跳过：{skip_count}首')
                            return
                        elif response:  # 覆盖
                            existing_poem.update(new_poem)
                            new_poems.remove(new_poem)
                            duplicate_count += 1
                            break
                        else:  # 跳过
                            new_poems.remove(new_poem)
                            skip_count += 1
                            break
            
            # 添加非重复诗词
            self.poems_data.extend(new_poems)
            
            # 保存到文件
            with open('poems.json', 'w', encoding='utf-8') as f:
                json.dump({'poems': self.poems_data}, f, ensure_ascii=False, indent=4)
                
            # 刷新显示
            self.search_poems()
            messagebox.showinfo('成功', 
                f'导入完成！\n成功导入：{len(new_poems)}首\n覆盖：{duplicate_count}首\n跳过：{skip_count}首')
            
        except Exception as e:
            messagebox.showerror('错误', f'导入失败：{str(e)}')

    def export_poems(self):
        try:
            # 打开文件保存对话框
            file_path = filedialog.asksaveasfilename(
                title='选择导出位置',
                defaultextension='.xlsx',  # 修改默认扩展名为xlsx
                filetypes=[
                    ('Excel files', '*.xlsx'),
                    ('JSON files', '*.json'),
                    ('CSV files', '*.csv')
                ]
            )
            
            if not file_path:
                return
                
            if file_path.lower().endswith('.json'):
                # 保存为JSON文件
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump({'poems': self.poems_data}, f, ensure_ascii=False, indent=4)
                
            elif file_path.lower().endswith('.xlsx'):
                try:
                    import openpyxl
                except ImportError:
                    messagebox.showerror('错误', '未找到openpyxl模块，无法导出为Excel文件。\n请使用以下命令安装：\npip install openpyxl')
                    return
                    
                # 创建工作簿和工作表
                wb = openpyxl.Workbook()
                ws = wb.active
                
                # 写入表头
                headers = ['title', 'author', 'dynasty', 'content', 'content_pinyin',
                          'translation', 'note', 'appreciation', 'author_intro', 'title_pinyin']
                ws.append(headers)
                
                # 写入数据
                for poem in self.poems_data:
                    row = []
                    for header in headers:
                        value = poem.get(header, '')
                        if header in ['content', 'content_pinyin']:
                            # 将列表转换为字符串
                            value = '|'.join(value) if isinstance(value, list) else value
                        row.append(value)
                    ws.append(row)
                
                # 调整列宽
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # 保存Excel文件
                wb.save(file_path)
                
            else:  # CSV文件
                with open(file_path, 'w', encoding='utf-8', newline='') as f:
                    # 定义CSV文件的字段
                    fieldnames = ['title', 'author', 'dynasty', 'content', 'content_pinyin',
                                'translation', 'note', 'appreciation', 'author_intro', 'title_pinyin']
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    # 写入诗词数据
                    for poem in self.poems_data:
                        row = poem.copy()
                        # 将列表转换为字符串
                        if isinstance(row.get('content'), list):
                            row['content'] = '|'.join(row['content'])
                        if isinstance(row.get('content_pinyin'), list):
                            row['content_pinyin'] = '|'.join(row['content_pinyin'])
                        writer.writerow(row)
                    
            messagebox.showinfo('成功', '导出成功！')
            
        except Exception as e:
            messagebox.showerror('错误', f'导出失败：{str(e)}')

    def show_help_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title('帮助')
        dialog.geometry('600x400')
        
        text = tk.Text(dialog, wrap=tk.WORD, padx=10, pady=10)
        text.pack(fill=tk.BOTH, expand=True)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(dialog, orient=tk.VERTICAL, command=text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text.configure(yscrollcommand=scrollbar.set)
        
        text.insert('1.0', help_text)
        text.config(state='disabled')
        
        # 设置对话框为模态
        dialog.transient(self.root)
        dialog.grab_set()

    def show_about_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title('关于')
        dialog.geometry('600x400')
        
        text = tk.Text(dialog, wrap=tk.WORD, padx=10, pady=10)
        text.pack(fill=tk.BOTH, expand=True)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(dialog, orient=tk.VERTICAL, command=text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text.configure(yscrollcommand=scrollbar.set)
        
        text.insert('1.0', about_text)
        text.config(state='disabled')
        
        # 设置对话框为模态
        dialog.transient(self.root)
        dialog.grab_set()

    def load_favorites(self):
        try:
            with open('favorites.json', 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            return []
        except Exception as e:
            print(f"加载收藏列表失败: {str(e)}")
            return []

    def save_favorites(self):
        try:
            with open('favorites.json', 'w', encoding='utf-8') as f:
                json.dump(self.favorites, f, ensure_ascii=False, indent=4)
        except Exception as e:
            messagebox.showerror('错误', f'保存收藏列表失败: {str(e)}')

    def toggle_favorite(self):
        selection = self.poem_tree.selection()
        if not selection:
            messagebox.showinfo('提示', '请先选择一首诗词')
            return
        
        item = selection[0]
        values = self.poem_tree.item(item)['values']
        title = values[0]
        author = values[1]
        
        # 创建唯一标识
        poem_id = f"{title}_{author}"
        
        if poem_id in self.favorites:
            # 取消收藏
            self.favorites.remove(poem_id)
            messagebox.showinfo('提示', f'已取消收藏《{title}》')
            self.favorite_btn.config(text='收藏')
        else:
            # 添加收藏
            self.favorites.append(poem_id)
            messagebox.showinfo('提示', f'已收藏《{title}》')
            self.favorite_btn.config(text='取消收藏')
        
        self.save_favorites()

    def show_favorites(self):
        """显示收藏的诗词"""
        # 清空现有列表
        for item in self.poem_tree.get_children():
            self.poem_tree.delete(item)
        
        # 显示收藏的诗词
        favorite_poems = []
        for poem in self.poems_data:
            poem_id = f"{poem['title']}_{poem['author']}"
            if poem_id in self.favorites:
                favorite_poems.append(poem)
        
        # 按标题排序
        favorite_poems.sort(key=lambda x: x['title'])
        
        # 显示收藏的诗词
        for poem in favorite_poems:
            self.poem_tree.insert('', 'end', values=(poem['title'], poem['author'], poem['dynasty']))
        
        # 仅在收藏夹为空时提示
        if not favorite_poems:
            messagebox.showinfo('提示', '收藏夹为空')

    def show_all_poems(self):
        """显示所有诗词"""
        # 清空搜索条件
        self.title_search.delete(0, tk.END)
        self.author_search.delete(0, tk.END)
        self.dynasty_search.delete(0, tk.END)
        
        # 显示所有诗词
        self.search_poems()

    def batch_delete_poems(self):
        # 获取选中的诗词
        selected_items = self.poem_tree.selection()
        if not selected_items:
            messagebox.showinfo('提示', '请先选择要删除的诗词')
            return
            
        # 获取选中诗词的标题和作者
        poems_to_delete = []
        for item in selected_items:
            values = self.poem_tree.item(item)['values']
            poems_to_delete.append(f'《{values[0]}》({values[1]})')
            
        # 显示确认对话框
        confirm = messagebox.askyesno('确认删除', 
            f'确定要删除以下{len(poems_to_delete)}首诗词吗？\n\n' + 
            '\n'.join(poems_to_delete))
            
        if confirm:
            # 执行删除操作
            for item in selected_items:
                values = self.poem_tree.item(item)['values']
                title = values[0]
                author = values[1]
                
                # 从数据中删除
                self.poems_data = [poem for poem in self.poems_data 
                    if not (poem['title'] == title and poem['author'] == author)]
                
            # 保存到文件
            with open('poems.json', 'w', encoding='utf-8') as f:
                json.dump({'poems': self.poems_data}, f, ensure_ascii=False, indent=4)
                
            # 刷新显示
            self.search_poems()
            messagebox.showinfo('成功', f'已删除{len(selected_items)}首诗词')
        else:
            # 取消选中状态
            self.poem_tree.selection_remove(*selected_items)

    def sort_tree(self, col):
        """按列排序树形视图"""
        # 如果是同一列，更新点击次数和排序方向
        if self.sort_column == col:
            self.sort_count += 1
            if self.sort_count > 2:  # 第三次点击
                self.sort_column = None
                self.sort_reverse = False
                self.sort_count = 0
                # 恢复原始顺序
                self.search_poems()  # 重新加载数据
                # 清除所有列标题的箭头
                for col_name in ('title', 'author', 'dynasty'):
                    text = self.poem_tree.heading(col_name)['text'].rstrip('↑↓')
                    self.poem_tree.heading(col_name, text=text)
                return
            else:
                self.sort_reverse = not self.sort_reverse
        else:
            # 切换到新列时重置状态
            self.sort_column = col
            self.sort_reverse = False
            self.sort_count = 1
        
        # 获取所有项目
        items = [(self.poem_tree.set(item, col), item) for item in self.poem_tree.get_children('')]
        
        # 使用拼音排序
        def get_pinyin(text):
            from pypinyin import pinyin, Style
            if not text:
                return ''
            # 获取拼音，转换为小写用于排序
            py_list = pinyin(text, style=Style.TONE3)
            return ''.join(p[0].lower() for p in py_list)
        
        # 按拼音排序
        items.sort(key=lambda x: get_pinyin(x[0]), reverse=self.sort_reverse)
        
        # 重新排列项目
        for index, (val, item) in enumerate(items):
            self.poem_tree.move(item, '', index)
        
        # 更新列标题显示排序方向
        for col_name in ('title', 'author', 'dynasty'):
            if col_name == col:
                direction = '↓' if self.sort_reverse else '↑'
                text = self.poem_tree.heading(col_name)['text'].rstrip('↑↓') + direction
            else:
                text = self.poem_tree.heading(col_name)['text'].rstrip('↑↓')
            self.poem_tree.heading(col_name, text=text)

if __name__ == '__main__':
    root = tk.Tk()
    app = PoemApp(root)
    root.mainloop()

    def batch_delete_poems(self):
        # 获取选中的诗词
        selected_items = self.poem_tree.selection()
        if not selected_items:
            messagebox.showinfo('提示', '请先选择要删除的诗词')
            return
            
        # 获取选中诗词的标题和作者
        poems_to_delete = []
        for item in selected_items:
            values = self.poem_tree.item(item)['values']
            poems_to_delete.append(f'《{values[0]}》({values[1]})')
            
        # 显示确认对话框
        confirm = messagebox.askyesno('确认删除', 
            f'确定要删除以下{len(poems_to_delete)}首诗词吗？\n\n' + 
            '\n'.join(poems_to_delete))
            
        if confirm:
            # 执行删除操作
            for item in selected_items:
                values = self.poem_tree.item(item)['values']
                title = values[0]
                author = values[1]
                
                # 从数据中删除
                self.poems_data = [poem for poem in self.poems_data 
                    if not (poem['title'] == title and poem['author'] == author)]
                
            # 保存到文件
            with open('poems.json', 'w', encoding='utf-8') as f:
                json.dump({'poems': self.poems_data}, f, ensure_ascii=False, indent=4)
                
            # 刷新显示
            self.search_poems()
            messagebox.showinfo('成功', f'已删除{len(selected_items)}首诗词')
        else:
            # 取消选中状态
            self.poem_tree.selection_remove(*selected_items)
        # 设置列标题的样式和背景色
        style = ttk.Style()
        style.configure('Treeview.Heading', background='#E8E8E8', relief='flat', font=('SimSun', 10))
        
        # 设置列标题
        self.poem_tree.heading('title', text='标题', command=lambda: self.sort_tree('title'))
        self.poem_tree.heading('author', text='作者', command=lambda: self.sort_tree('author'))
        self.poem_tree.heading('dynasty', text='朝代', command=lambda: self.sort_tree('dynasty'))
        
        # 修改排序状态记录
        self.sort_column = None  # 当前排序的列
        self.sort_reverse = False  # 是否逆序
        self.sort_count = 0  # 添加点击次数记录
        
        # 配置left_frame的网格权重，使poem_tree能够自动扩展
        self.left_frame.grid_columnconfigure(0, weight=1)
        self.left_frame.grid_rowconfigure(1, weight=1)

        # 设置列宽
        self.poem_tree.column('title', width=150)
        self.poem_tree.column('author', width=100)
        self.poem_tree.column('dynasty', width=40)

        # 绑定选择事件
        self.poem_tree.bind('<<TreeviewSelect>>', self.show_poem_details)

        # 添加滚动条
        tree_scroll = ttk.Scrollbar(self.left_frame, orient=tk.VERTICAL, command=self.poem_tree.yview)
        tree_scroll.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.poem_tree.configure(yscrollcommand=tree_scroll.set)

        # 创建右侧内容框架
        self.content_frame = ttk.Frame(self.right_frame)
        self.content_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.content_frame.grid_columnconfigure(0, weight=1)
        self.content_frame.grid_rowconfigure(1, weight=2)  # 诗词显示框占2份
        self.content_frame.grid_rowconfigure(2, weight=3)  # 注释内容显示框占3份

        # 创建右侧顶部按钮框架
        self.right_top_frame = ttk.Frame(self.content_frame)
        self.right_top_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        self.right_top_frame.grid_columnconfigure(0, weight=1)

        # 创建一个内部框架来容纳按钮，实现居中效果
        button_container = ttk.Frame(self.right_top_frame)
        button_container.grid(row=0, column=0)

        # 添加编辑、保存、取消、关闭、导入导出按钮到按钮容器中
        self.add_btn = ttk.Button(button_container, text='添加', command=self.add_poem)
        self.add_btn.pack(side=tk.LEFT, padx=5)
        self.favorite_btn = ttk.Button(button_container, text='收藏', command=self.toggle_favorite)
        self.favorite_btn.pack(side=tk.LEFT, padx=5)
        self.edit_btn = ttk.Button(button_container, text='编辑', command=self.edit_poem)
        self.edit_btn.pack(side=tk.LEFT, padx=5)
        self.save_btn = ttk.Button(button_container, text='保存', command=self.save_poem, state='disabled')
        self.save_btn.pack(side=tk.LEFT, padx=5)
        self.cancel_btn = ttk.Button(button_container, text='取消', command=self.cancel_edit, state='disabled')
        self.cancel_btn.pack(side=tk.LEFT, padx=5)
        self.import_btn = ttk.Button(button_container, text='导入', command=self.import_poems)
        self.import_btn.pack(side=tk.LEFT, padx=5)
        self.export_btn = ttk.Button(button_container, text='导出', command=self.export_poems)
        self.export_btn.pack(side=tk.LEFT, padx=5)
        self.close_btn = ttk.Button(button_container, text='退出', command=self.close_app)
        self.close_btn.pack(side=tk.LEFT, padx=5)

        # 修改诗词内容文本框的创建
        self.poem_content = tk.Text(self.content_frame, wrap=tk.WORD, font=('SimSun', 14), 
                                   relief=tk.SUNKEN, borderwidth=1)
        self.poem_content.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 2))
        
        # 添加滚动条
        content_scroll = ttk.Scrollbar(self.content_frame, orient=tk.VERTICAL, command=self.poem_content.yview)
        content_scroll.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.poem_content.configure(yscrollcommand=content_scroll.set)
        
        # 创建朗读控制面板
        self.read_control_frame = ttk.LabelFrame(self.content_frame, text='朗读控制', padding=5)
        self.read_control_frame.grid(row=1, column=2, sticky=(tk.N, tk.S, tk.E), padx=5, pady=5)
        
        # 添加音色选择
        ttk.Label(self.read_control_frame, text='音色:').pack(pady=(5,0))
        self.voice_var = tk.StringVar()
        self.voice_combo = ttk.Combobox(self.read_control_frame, textvariable=self.voice_var, state='readonly')
        self.voice_combo.pack(fill=tk.X, pady=1)
        
        # 添加朗读按钮
        self.read_btn = ttk.Button(self.read_control_frame, text='朗读', command=self.read_poem)
        self.read_btn.pack(fill=tk.X, pady=1)
        
        # 添加暂停/继续按钮
        self.pause_btn = ttk.Button(self.read_control_frame, text='暂停', command=self.pause_resume_reading, state='disabled')
        self.pause_btn.pack(fill=tk.X, pady=1)
        
        # 添加停止按钮
        self.stop_btn = ttk.Button(self.read_control_frame, text='停止', command=self.stop_reading, state='disabled')
        self.stop_btn.pack(fill=tk.X, pady=1)
        
        # 添加语速控制
        ttk.Label(self.read_control_frame, text='语速:').pack(pady=(10,0))
        self.rate_scale = ttk.Scale(self.read_control_frame, from_=50, to=300, orient=tk.HORIZONTAL)
        self.rate_scale.set(150)
        self.rate_scale.pack(fill=tk.X, pady=1)
        
        # 添加音量控制
        ttk.Label(self.read_control_frame, text='音量:').pack(pady=(10,0))
        self.volume_scale = ttk.Scale(self.read_control_frame, from_=0, to=1, orient=tk.HORIZONTAL)
        self.volume_scale.set(1.0)
        self.volume_scale.pack(fill=tk.X, pady=1)
        
        # 添加退出按钮到朗读控制区域下方
        self.close_btn = ttk.Button(self.read_control_frame, text='退出程序', command=self.close_app, width=8)
        self.close_btn.pack(fill=tk.X, pady=(10,1))
        
        # 创建诗词详情的框架
        self.detail_frame = ttk.Frame(self.content_frame)
        self.detail_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        self.detail_frame.grid_columnconfigure(0, weight=1)
        self.detail_frame.grid_rowconfigure(0, weight=1)
        
        # 创建诗词详情的文本框和标签
        info_frame = ttk.Frame(self.detail_frame)
        info_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))